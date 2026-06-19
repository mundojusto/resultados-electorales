#!/usr/bin/env python3
"""Procesador de resultados electorales oficiales para el partido M+J.

Lee un fichero XLSX oficial con los votos por municipio (Congreso, Senado,
Europeas o Cabildos), localiza la columna del partido "Por Un Mundo Más Justo"
(M+J / PUM+J / MUNDO+JUSTO) —o de la coalición en la que concurrió, como
"Existe" en las europeas— y guarda un JSON únicamente con esos resultados.

Uso:
    python procesar_resultados.py ENTRADA.xlsx [--salida DIR] [--borrar-entrada]

El fichero se busca de forma robusta: se detecta la fila de cabeceras por la
celda "Nombre de Comunidad", la fila de nombres completos justo encima, y la
fila de título (con "Resultados por") más arriba.
"""
from __future__ import annotations

import argparse
import json
import re
import sys
import unicodedata
from datetime import date
from pathlib import Path

import openpyxl

# --- Identificadores del partido M+J ------------------------------------------
# Se comparan sobre el texto normalizado (mayúsculas, sin acentos). Basta con
# que el nombre completo O las siglas coincidan con alguno de estos patrones.
#
# Se incluyen también las coaliciones en las que M+J ha concurrido sin marca
# propia en la papeleta: en las elecciones europeas M+J se presentó dentro de la
# coalición "Existe", de modo que en el XLSX oficial la columna aparece con el
# nombre/siglas de la coalición y no con los de M+J. En esos casos los votos
# recogidos son los de la candidatura conjunta (coalición), no solo los de M+J.
PATRONES_NOMBRE = [
    "POR UN MUNDO MAS JUSTO",
    "MUNDO MAS JUSTO",
    "MUNDO+JUSTO",
    "EXISTE",
]
PATRONES_SIGLAS = ["PUM+J", "M+J", "PUMJ", "MUNDO+JUSTO", "EXISTE"]

# En municipales, se pide limitar la búsqueda a estas marcas exactas.
PATRONES_MUNICIPALES_EXACTOS = {"PUM+J", "M+J", "MUNDO+JUSTO"}

# --- Cabeceras de metadatos (columnas no-partido) -----------------------------
META_COMUNIDAD = "Nombre de Comunidad"
COLUMNAS_META = {
    "Nombre de Comunidad": "comunidad",
    "Código de Provincia": "codigo_provincia",
    "Nombre de Provincia": "provincia",
    "Código de Municipio": "codigo_municipio",
    "Nombre de Municipio": "municipio",
    "Población": "poblacion",
    "Número de mesas": "mesas",
    "Total censo electoral": "censo",
    "Total votantes": "votantes",
    "Votos válidos": "votos_validos",
    "Votos a candidaturas": "votos_candidaturas",
    "Votos en blanco": "votos_blanco",
    "Votos nulos": "votos_nulos",
}

MESES = {
    "enero": 1, "febrero": 2, "marzo": 3, "abril": 4, "mayo": 5, "junio": 6,
    "julio": 7, "agosto": 8, "septiembre": 9, "octubre": 10, "noviembre": 11,
    "diciembre": 12,
}


def normaliza(texto) -> str:
    """Mayúsculas, sin acentos y con espacios colapsados."""
    if texto is None:
        return ""
    s = unicodedata.normalize("NFKD", str(texto))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", s).strip().upper()


def detecta_filas(ws):
    """Devuelve (fila_titulo, fila_nombres, fila_cabecera) en índice 1-based."""
    fila_cabecera = None
    fila_titulo = None
    meta = normaliza(META_COMUNIDAD)

    for i, fila in enumerate(ws.iter_rows(min_row=1, max_row=30, values_only=True), start=1):
        celdas = tuple(fila or ())

        # En algunos XLSX la columna de metadatos empieza en la columna B,
        # así que buscamos la cabecera en toda la fila y no solo en la primera celda.
        if any(normaliza(celda) == meta for celda in celdas):
            fila_cabecera = i
            break

        if any("RESULTADOS POR" in normaliza(celda) for celda in celdas):
            fila_titulo = i

    if fila_cabecera is None:
        raise ValueError("No se encontró la fila de cabeceras ('Nombre de Comunidad').")
    return fila_titulo, fila_cabecera - 1, fila_cabecera


def parsea_titulo(texto):
    """De 'Congreso | Julio 2023 | Resultados por municipio' extrae metadatos."""
    partes = [p.strip() for p in str(texto).split("|")] if texto else []
    tipo = partes[0] if partes else None
    periodo = partes[1] if len(partes) > 1 else None
    anio = mes = None
    if periodo:
        m = re.search(r"(\d{4})", periodo)
        if m:
            anio = int(m.group(1))
        for nombre, num in MESES.items():
            if nombre in normaliza(periodo).lower():
                mes = num
                break
    return tipo, periodo, anio, mes


def localiza_columnas(tipo, fila_nombres, fila_cabecera):
    """Mapea columnas de metadatos y localiza la columna del partido M+J.

    Devuelve (indices_meta: dict, idx_partido: int, nombre_partido, siglas_partido).
    """
    indices_meta = {}
    norm_meta = {normaliza(k): v for k, v in COLUMNAS_META.items()}
    idx_partido = None
    nombre_partido = siglas_partido = None

    tipo_normalizado = normaliza(tipo)
    es_municipales = "MUNICIPALES" in tipo_normalizado
    patrones_siglas_norm = {normaliza(s) for s in PATRONES_SIGLAS}
    patrones_municipales_norm = {normaliza(s) for s in PATRONES_MUNICIPALES_EXACTOS}

    for idx, siglas in enumerate(fila_cabecera):
        clave = norm_meta.get(normaliza(siglas))
        if clave:
            indices_meta[clave] = idx
            continue
        nombre = fila_nombres[idx] if idx < len(fila_nombres) else None
        n_nombre, n_siglas = normaliza(nombre), normaliza(siglas)
        if not n_nombre and not n_siglas:
            continue
        if es_municipales:
            es_mj = n_siglas in patrones_municipales_norm or n_nombre in patrones_municipales_norm
        else:
            es_mj = any(p in n_nombre for p in PATRONES_NOMBRE) or n_siglas in patrones_siglas_norm
        if es_mj:
            idx_partido = idx
            nombre_partido = str(nombre).strip() if nombre else str(siglas).strip()
            siglas_partido = str(siglas).strip() if siglas else None

    faltan = {"comunidad", "provincia", "municipio", "votos_validos"} - indices_meta.keys()
    if faltan:
        raise ValueError(f"Faltan columnas de metadatos esperadas: {sorted(faltan)}")
    # idx_partido puede ser None si M+J no concurrió a esta elección; lo gestiona procesa().
    return indices_meta, idx_partido, nombre_partido, siglas_partido


class PartidoNoPresente(Exception):
    """M+J no concurrió a esta elección (no hay columna del partido)."""

    def __init__(self, tipo, periodo):
        self.tipo = tipo
        self.periodo = periodo
        super().__init__(f"M+J no se presentó en {tipo} {periodo}")


def a_int(valor):
    try:
        return int(valor)
    except (TypeError, ValueError):
        return 0


def slug(texto):
    return re.sub(r"[^a-z0-9]+", "-", normaliza(texto).lower()).strip("-")


def procesa(entrada: Path):
    wb = openpyxl.load_workbook(entrada, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    fila_titulo, fila_nombres_idx, fila_cabecera_idx = detecta_filas(ws)
    cabeceras = list(ws.iter_rows(min_row=1, max_row=fila_cabecera_idx, values_only=True))
    fila_nombres = cabeceras[fila_nombres_idx - 1] if fila_nombres_idx >= 1 else ()
    fila_cabecera = cabeceras[fila_cabecera_idx - 1]

    titulo = None
    if fila_titulo:
        fila_t = cabeceras[fila_titulo - 1]
        titulo = next((celda for celda in fila_t if celda not in (None, "")), None)

    tipo, periodo, anio, mes = parsea_titulo(titulo)
    meta, idx_p, nombre_p, siglas_p = localiza_columnas(tipo, fila_nombres, fila_cabecera)
    if idx_p is None:
        raise PartidoNoPresente(tipo, periodo)

    resultados = []
    tot_partido = tot_validos = tot_candidaturas = 0
    for fila in ws.iter_rows(min_row=fila_cabecera_idx + 1, values_only=True):
        if fila[meta["municipio"]] in (None, ""):
            continue
        cod_prov = (a_int(fila[meta.get("codigo_provincia", -1)])
                    if "codigo_provincia" in meta else None)
        cod_mun = (a_int(fila[meta.get("codigo_municipio", -1)])
                   if "codigo_municipio" in meta else None)
        votos = a_int(fila[idx_p])
        validos = a_int(fila[meta["votos_validos"]])
        candidaturas = a_int(fila[meta.get("votos_candidaturas", meta["votos_validos"])])
        tot_partido += votos
        tot_validos += validos
        tot_candidaturas += candidaturas
        registro = {
            "comunidad": fila[meta["comunidad"]],
            "codigo_provincia": cod_prov,
            "provincia": fila[meta["provincia"]],
            "codigo_municipio": cod_mun,
            "municipio": fila[meta["municipio"]],
            "codigo_ine": f"{cod_prov:02d}{cod_mun:03d}" if cod_prov and cod_mun else None,
            "censo": a_int(fila[meta["censo"]]) if "censo" in meta else None,
            "votantes": a_int(fila[meta["votantes"]]) if "votantes" in meta else None,
            "votos_validos": validos,
            "votos_candidaturas": candidaturas,
            "votos_partido": votos,
        }
        resultados.append(registro)

    return {
        "eleccion": {
            "tipo": tipo,
            "periodo": periodo,
            "anio": anio,
            "mes": mes,
            "ambito": "municipio",
            "fuente_archivo": entrada.name,
            "fecha_procesado": date.today().isoformat(),
        },
        "partido": {"nombre": nombre_p, "siglas": siglas_p},
        "totales": {
            "votos_partido": tot_partido,
            "votos_validos": tot_validos,
            "votos_candidaturas": tot_candidaturas,
            "porcentaje_validos": round(100 * tot_partido / tot_validos, 4) if tot_validos else 0,
            "municipios": len(resultados),
        },
        "resultados": resultados,
    }


def nombre_salida(datos):
    e = datos["eleccion"]
    partes = [slug(e["tipo"]) or "eleccion"]
    if e["anio"]:
        partes.append(f"{e['anio']}{e['mes']:02d}" if e["mes"] else str(e["anio"]))
    return "_".join(partes) + ".json"


def main(argv=None):
    p = argparse.ArgumentParser(
        description="Procesa un XLSX oficial y extrae los resultados de M+J.")
    p.add_argument("entrada", type=Path, help="Fichero XLSX oficial a procesar")
    p.add_argument("--salida", type=Path, default=Path("resultados-oficiales-procesados"),
                   help="Directorio de salida para el JSON "
                        "(por defecto: resultados-oficiales-procesados)")
    p.add_argument("--borrar-entrada", action="store_true",
                   help="Borra el fichero de entrada tras procesarlo correctamente")
    args = p.parse_args(argv)

    if not args.entrada.exists():
        print(f"ERROR: no existe el fichero {args.entrada}", file=sys.stderr)
        return 1

    try:
        datos = procesa(args.entrada)
    except PartidoNoPresente as e:
        # No es un error: M+J no concurrió a esa elección. Se omite sin generar JSON.
        print(f"OMITIDO  {args.entrada.name}: {e}. No se genera JSON.")
        if args.borrar_entrada:
            args.entrada.unlink()
            print(f"    fichero de entrada borrado: {args.entrada}")
        return 0

    args.salida.mkdir(parents=True, exist_ok=True)
    destino = args.salida / nombre_salida(datos)
    with destino.open("w", encoding="utf-8") as f:
        json.dump(datos, f, ensure_ascii=False, indent=2)

    t = datos["totales"]
    print(f"OK  {args.entrada.name} -> {destino}")
    print(f"    {datos['eleccion']['tipo']} {datos['eleccion']['periodo']} | "
          f"partido: {datos['partido']['nombre']} ({datos['partido']['siglas']})")
    print(f"    {t['municipios']} municipios | {t['votos_partido']} votos M+J | "
          f"{t['porcentaje_validos']}% s/válidos")

    if args.borrar_entrada:
        args.entrada.unlink()
        print(f"    fichero de entrada borrado: {args.entrada}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
